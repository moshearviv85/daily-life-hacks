"""
Build pipeline-data/content-batch-60.json and pipeline-data/content-inventory-2026.xlsx

Run from repo root: python scripts/build_content_batch_60_workbook.py
"""
from __future__ import annotations

import csv
import json
import re
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
ART_DIR = ROOT / "src" / "data" / "articles"
TOPICS_FILE = ROOT / "pipeline-data" / "topics-to-write.md"
PUBLER = ROOT / "pipeline-data" / "pins-publer-final.csv"
OUT_JSON = ROOT / "pipeline-data" / "content-batch-60.json"
OUT_XLSX = ROOT / "pipeline-data" / "content-inventory-2026.xlsx"

# Reference "today" for Publer published vs pending (align with project calendar)
TODAY = datetime(2026, 4, 3, 12, 0, tzinfo=timezone.utc)

# 60 new topics: 20 recipes, 20 nutrition, 20 tips — diverse angles (not fiber-only)
RECIPES = [
    ("mediterranean-baked-cod-tomatoes-olives", "Mediterranean baked cod with tomatoes and olives", "Mediterranean"),
    ("thai-inspired-peanut-tofu-noodle-bowl", "Thai-inspired peanut tofu noodle bowl", "Asian"),
    ("miso-soup-tofu-wakame-weeknight", "Miso soup with tofu and wakame for busy nights", "Japanese"),
    ("easy-red-lentil-dal-coconut", "Easy red lentil dal with coconut", "Indian"),
    ("korean-inspired-beef-vegetable-rice-bowl", "Korean-inspired beef and vegetable rice bowl", "Korean"),
    ("vegetarian-stuffed-peppers-mediterranean", "Vegetarian stuffed peppers Mediterranean style", "Mediterranean"),
    ("chickpea-shawarma-sheet-pan-dinner", "Chickpea shawarma sheet pan dinner", "Middle Eastern"),
    ("vietnamese-inspired-chicken-lettuce-cups", "Vietnamese-inspired chicken lettuce cups", "Vietnamese"),
    ("cashew-cream-pasta-vegan-weeknight", "Cashew cream pasta vegan weeknight", "Vegan"),
    ("greek-yogurt-marinated-chicken-skewers", "Greek yogurt marinated chicken skewers", "Mediterranean"),
    ("air-fryer-salmon-citrus-herbs", "Air fryer salmon with citrus and herbs", "Weeknight"),
    ("lentil-mushroom-bolognese-vegetarian", "Lentil mushroom bolognese vegetarian", "Italian"),
    ("cauliflower-fried-rice-with-eggs", "Cauliflower fried rice with eggs", "Asian"),
    ("teriyaki-tempeh-broccoli-rice-bowl", "Teriyaki tempeh and broccoli rice bowl", "Asian"),
    ("mexican-spiced-black-bean-soup", "Mexican spiced black bean soup", "Mexican"),
    ("rosemary-white-bean-tomato-soup", "Rosemary white bean and tomato soup", "Mediterranean"),
    ("banana-oat-pancakes-no-flour", "Banana oat pancakes without flour", "Breakfast"),
    ("spinach-feta-egg-bites-meal-prep", "Spinach feta egg bites for meal prep", "Breakfast"),
    ("pesto-zucchini-noodles-white-beans", "Pesto zucchini noodles with white beans", "Mediterranean"),
    ("harissa-roasted-carrots-chickpeas", "Harissa roasted carrots and chickpeas", "North African"),
]

NUTRITION = [
    ("mediterranean-diet-staples-beginner-list", "Mediterranean diet staples list for beginners", "Mediterranean"),
    ("plant-based-iron-sources-without-meat", "Plant-based iron sources without meat", "Plant-based"),
    ("sodium-smart-swaps-home-cooking", "Sodium-smart swaps for home cooking", "Heart-health"),
    ("balanced-salad-that-actually-fills-you-up", "How to build a balanced salad that fills you up", "Practical"),
    ("snack-label-reading-added-sugar-guide", "Snack label reading for added sugar", "Labels"),
    ("protein-at-breakfast-busy-mornings", "Protein at breakfast ideas for busy mornings", "Protein"),
    ("whole-grains-vs-refined-what-changes", "Whole grains vs refined grains what changes", "Basics"),
    ("hydration-and-salty-meals-practical-tips", "Hydration and salty meals practical tips", "Hydration"),
    ("anti-inflammatory-spices-everyday-cooking", "Anti-inflammatory spices for everyday cooking", "Spices"),
    ("budget-protein-sources-that-last", "Budget protein sources that last in the pantry", "Budget"),
    ("cooking-oils-which-fat-for-what", "Cooking oils which fat for what heat", "Fats"),
    ("beans-without-gut-drama-practical-tips", "How to eat more beans without gut drama", "Practical"),
    ("prebiotic-foods-beyond-the-buzzwords", "Prebiotic foods beyond the buzzwords", "Gut"),
    ("electrolyte-foods-without-sports-drinks", "Electrolyte foods without sports drinks", "Hydration"),
    ("meal-timing-and-energy-what-people-actually-try", "Meal timing and energy what people actually try", "Energy"),
    ("night-shift-snacking-strategies", "Night shift snacking strategies that are realistic", "Shift work"),
    ("frozen-vegetables-that-taste-good", "How to choose frozen vegetables that taste good", "Shopping"),
    ("plant-milks-for-coffee-what-to-expect", "Plant milks for coffee what to expect", "Shopping"),
    ("low-sugar-breakfast-cereal-picks", "Low sugar breakfast cereal picks without the lecture", "Breakfast"),
    ("asian-inspired-balanced-plate-ideas", "Asian-inspired balanced plate ideas for home cooks", "Asian"),
]

TIPS = [
    ("how-to-dice-onion-faster-with-less-drama", "How to dice an onion faster with less drama", "Knife skills"),
    ("how-to-keep-basil-fresh-longer", "How to keep basil fresh longer", "Herbs"),
    ("how-to-clean-blender-garlic-smell", "How to clean a blender that smells like garlic", "Cleanup"),
    ("how-to-revive-limp-celery", "How to revive limp celery", "Storage"),
    ("how-to-prep-garlic-ahead-of-time", "How to prep garlic ahead of time", "Prep"),
    ("how-to-store-cut-onions-with-less-smell", "How to store cut onions with less smell", "Storage"),
    ("how-to-test-pan-heat-before-searing", "How to test pan heat before searing", "Cooking"),
    ("how-to-line-sheet-pan-for-easy-cleanup", "How to line a sheet pan for easy cleanup", "Cleanup"),
    ("how-to-keep-salad-greens-crispy", "How to keep salad greens crispy in the fridge", "Storage"),
    ("how-to-salt-pasta-water-properly", "How to salt pasta water properly", "Pasta"),
    ("how-to-thaw-chicken-safely", "How to thaw chicken safely", "Food safety"),
    ("how-to-pack-salad-for-work-not-soggy", "How to pack salad for work without sogginess", "Meal prep"),
    ("how-to-cool-rice-for-fried-rice", "How to cool rice for better fried rice", "Rice"),
    ("how-to-stretch-leftover-roast-chicken", "How to stretch leftover roast chicken", "Leftovers"),
    ("how-to-organize-spices-on-a-budget", "How to organize spices on a budget", "Organization"),
    ("how-to-pick-ripe-melon-at-the-store", "How to pick a ripe melon at the store", "Shopping"),
    ("how-to-store-citrus-so-it-lasts", "How to store citrus so it lasts", "Storage"),
    ("how-to-reduce-smoke-when-searing-meat", "How to reduce smoke when searing meat", "Technique"),
    ("how-to-soften-butter-quickly-for-baking", "How to soften butter quickly for baking", "Baking"),
    ("how-to-rinse-rice-for-fluffier-grains", "How to rinse rice for fluffier grains", "Rice"),
]


def load_yaml_frontmatter(path: Path) -> dict:
    text = path.read_text(encoding="utf-8")
    parts = text.split("---", 2)
    if len(parts) < 3:
        return {}
    try:
        import yaml

        return yaml.safe_load(parts[1]) or {}
    except Exception:
        return {}


def existing_slugs() -> set[str]:
    return {p.stem for p in ART_DIR.glob("*.md")}


def topics_backlog_slugs() -> set[str]:
    text = TOPICS_FILE.read_text(encoding="utf-8")
    return set(re.findall(r"\|\s*\d+\s*\|\s*\w+\s*\|\s*[^|]+\|\s*([a-z0-9-]+)\s*\|", text))


def max_publish_at_iso(articles: list[dict]) -> str | None:
    best = None
    for a in articles:
        p = a.get("publishAt") or ""
        if not p:
            continue
        s = p.replace("Z", "+00:00")
        try:
            d = datetime.fromisoformat(s)
            if best is None or d > best:
                best = d
        except ValueError:
            continue
    return best.isoformat() if best else None


def build_batch_json(
    existing: set[str], backlog: set[str], start: date
) -> tuple[list[dict], list[str]]:
    errors: list[str] = []
    # Round-robin: recipes, nutrition, tips
    triples = list(zip(RECIPES, NUTRITION, TIPS))
    out: list[dict] = []
    for i, (r, n, t) in enumerate(triples):
        day = start + timedelta(days=i * 3)
        for j, (slug, title, theme) in enumerate(
            [
                (r[0], r[1], r[2]),
                (n[0], n[1], n[2]),
                (t[0], t[1], t[2]),
            ]
        ):
            d = day + timedelta(days=j)
            cat = ["recipes", "nutrition", "tips"][j]
            if slug in existing:
                errors.append(f"DUPLICATE existing site slug: {slug}")
            if slug in backlog:
                errors.append(f"DUPLICATE backlog topics-to-write slug: {slug}")
            out.append(
                {
                    "order": len(out) + 1,
                    "slug": slug,
                    "category": cat,
                    "working_title": title,
                    "theme": theme,
                    "publishAt": f"{d.isoformat()}T00:00:00.000Z",
                    "main_image": f"/images/{slug}-main.jpg",
                    "pin_images": [
                        f"/images/pins/{slug}_v1.jpg",
                        f"/images/pins/{slug}_v2.jpg",
                        f"/images/pins/{slug}_v3.jpg",
                        f"/images/pins/{slug}_v4.jpg",
                    ],
                }
            )
    return out, errors


def load_articles_inventory() -> list[dict]:
    rows = []
    for p in sorted(ART_DIR.glob("*.md")):
        data = load_yaml_frontmatter(p)
        pub = data.get("publishAt")
        d = data.get("date")
        rows.append(
            {
                "slug": p.stem,
                "category": data.get("category") or "",
                "title": data.get("title") or "",
                "date": str(d) if d else "",
                "publishAt": str(pub) if pub else "",
            }
        )
    return rows


def effective_release(row: dict) -> datetime | None:
    if row.get("publishAt"):
        s = row["publishAt"].replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(s)
        except ValueError:
            pass
    if row.get("date"):
        try:
            return datetime.fromisoformat(str(row["date"]) + "T12:00:00+00:00")
        except ValueError:
            pass
    return None


def parse_publer_rows() -> list[dict]:
    rows = []
    with PUBLER.open(encoding="utf-8", newline="") as f:
        r = csv.reader(f)
        header = next(r, None)
        for line in r:
            if len(line) < 4:
                continue
            dt_s, text, links, media = line[0], line[1], line[2], line[3]
            title = line[4] if len(line) > 4 else ""
            board = line[8] if len(line) > 8 else ""
            try:
                dt = datetime.strptime(dt_s.strip(), "%Y-%m-%d %H:%M")
                dt = dt.replace(tzinfo=timezone.utc)
            except ValueError:
                dt = None
            url = links.strip().split(",")[0].strip() if links else ""
            path_slug = ""
            if "daily-life-hacks.com/" in url:
                path_slug = url.split("daily-life-hacks.com/", 1)[-1].split("?")[0].strip("/").split("/")[-1]
            status = ""
            if dt:
                status = "Published (Publer date in past)" if dt < TODAY else "Scheduled / pending"
            rows.append(
                {
                    "publer_datetime_utc": dt.isoformat() if dt else "",
                    "pin_title": title,
                    "destination_url": url,
                    "path_slug": path_slug,
                    "board": board,
                    "media_url": media.strip().split(",")[0].strip() if media else "",
                    "publer_status": status,
                }
            )
    return rows


def main() -> None:
    existing = existing_slugs()
    backlog = topics_backlog_slugs()
    inv = load_articles_inventory()

    max_pub = None
    for row in inv:
        if row.get("publishAt"):
            s = row["publishAt"].replace("Z", "+00:00")
            try:
                d = datetime.fromisoformat(s)
                if max_pub is None or d > max_pub:
                    max_pub = d
            except ValueError:
                pass

    if max_pub:
        start_day = (max_pub.date() + timedelta(days=1)) if max_pub.tzinfo else (max_pub.date() + timedelta(days=1))
    else:
        start_day = date(2026, 4, 12)

    batch, errs = build_batch_json(existing, backlog, start_day)
    if errs:
        raise SystemExit("Validation failed:\n" + "\n".join(errs))

    OUT_JSON.write_text(json.dumps(batch, indent=2), encoding="utf-8")

    wb = Workbook()
    # Sheet 1: Articles on site
    ws1 = wb.active
    ws1.title = "Articles_On_Site"
    h1 = [
        "slug",
        "category",
        "title",
        "date",
        "publishAt",
        "effective_release_utc",
        "site_release_status",
        "file_on_disk",
    ]
    ws1.append(h1)
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    for row in sorted(inv, key=lambda x: x["slug"]):
        eff = effective_release(row)
        eff_s = eff.isoformat() if eff else ""
        if eff:
            st = "Released (as of 2026-04-03 UTC ref)" if eff <= TODAY else "Scheduled future release"
        else:
            st = "Unknown"
        ws1.append(
            [
                row["slug"],
                row["category"],
                row["title"],
                row["date"],
                row["publishAt"],
                eff_s,
                st,
                "yes",
            ]
        )

    # Sheet 2: Pins Publer runway
    ws2 = wb.create_sheet("Pins_Publer_Runway")
    publer = parse_publer_rows()
    h2 = list(publer[0].keys()) if publer else ["empty"]
    ws2.append(h2)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for pr in publer:
        ws2.append([pr[k] for k in h2])

    # Sheet 3: Pin summary
    ws3 = wb.create_sheet("Pins_Summary")
    ws3.append(["Metric", "Value"])
    ws3["A1"].font = Font(bold=True)
    ws3["B1"].font = Font(bold=True)
    pub_past = sum(1 for p in publer if p.get("publer_status") == "Published (Publer date in past)")
    pub_future = sum(1 for p in publer if p.get("publer_status") == "Scheduled / pending")
    ws3.append(["Total Publer rows (pins-publer-final.csv)", len(publer)])
    ws3.append(["Pins with Publer datetime in the past (treated as published)", pub_past])
    ws3.append(["Pins with Publer datetime today or future (pending)", pub_future])
    ws3.append(["Reference datetime for comparison", TODAY.isoformat()])
    ws3.append(["Source file", str(PUBLER.relative_to(ROOT))])

    # Sheet 4: New batch 60
    ws4 = wb.create_sheet("New_Batch_60")
    keys = list(batch[0].keys()) if batch else []
    ws4.append(keys)
    for cell in ws4[1]:
        cell.font = Font(bold=True)
    for item in batch:
        ws4.append([json.dumps(item[k]) if isinstance(item[k], list) else item[k] for k in keys])

    # Sheet 5: Router note
    ws5 = wb.create_sheet("Pinterest_URL_Notes")
    notes = [
        "Query parameters on Pinterest links do not change the pathname.",
        "The Smart Router (functions/[[path]].js) uses url.pathname only for KV lookup and -v{n} fallback.",
        "UTM and other query strings are preserved in analytics logging but do not break routing.",
        "Ensure ROUTES_KV entries exist for keyword slugs so /keyword-slug proxies to base_slug.",
    ]
    for i, n in enumerate(notes, 1):
        ws5.cell(row=i, column=1, value=n)

    for ws in (ws1, ws2, ws4):
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    wb.save(OUT_XLSX)

    print("Wrote", OUT_JSON.relative_to(ROOT))
    print("Wrote", OUT_XLSX.relative_to(ROOT))
    print("Batch start date", start_day.isoformat(), "max prior publishAt", max_publish_at_iso(inv))
    print("Batch rows", len(batch))


if __name__ == "__main__":
    main()
